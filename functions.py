#!/usr/bin/env python

from googletrans import Translator
import os
import random
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

translator = Translator()


def introStuff(fileName, categoryName):
    PATH = "new_files"
    if not os.path.exists(PATH):
        os.makedirs(PATH)
    num = random.randrange(1, 100)
    if not categoryName:
        w = open('new_files/{}_pleco.txt'.format(fileName), 'w')
        w.write("//Lazy Category Namer {}\n".format(num))
    else:
        w = open('new_files/{}_pleco.txt'.format(fileName), 'w')
        w.write("//{}\n".format(categoryName))


def make_xls_worksheet(werdz, tab_name):
    filepath = 'new_files/yoyo_生词.xlsx'
    if not os.path.exists(filepath):
        wb = Workbook()
    else:
        wb = load_workbook(filename=filepath)
    ws = wb.create_sheet('又o又o Spreadsheet', 0)
    ws.title = tab_name

    # title row
    title_row = [['A1', 'Simplified'], ['B1', 'Traditional'],
                 ['C1', 'Pinyin'], ['D1', 'English Definition']]
    for title in title_row:
        ws[title[0]] = title[1]
        ws[title[0]].font = Font(bold=True)

    rownum = 2

    # data stuffs
    for werd in werdz:
        transSF = werd.strip()
        transEN = translator.translate(transSF, dest='en')
        transLF = translator.translate(transSF, dest='zh-tw')
        transPY = translator.translate(transSF, dest='zh-cn')

        ws.cell(row=rownum, column=1, value=transSF)
        ws.cell(row=rownum, column=2, value=transLF.text)
        ws.cell(row=rownum, column=3, value=transPY.pronunciation)
        ws.cell(row=rownum, column=4, value=transEN.text)
        rownum += 1

    wb.save(filepath)


def zhengLongShuo(werdz):
    split = werdz.split(',')
    return split

# 郑隆比周星大好得多。


def newLineSplit(werdz):
    split = werdz.split('\n')
    return split


def getTheWerdz(werdz, fName):
    for werd in werdz:
        transSF = werd.strip()
        transEN = translator.translate(transSF, dest='en')
        transLF = translator.translate(transSF, dest='zh-tw')
        transPY = translator.translate(transSF, dest='zh-cn')

        f = open('new_files/{}_pleco.txt'.format(fName), 'a')
        f.write("{}[{}]\t{}\t{}\n".format(
            transSF, transLF.text, transPY.pronunciation, transEN.text))
